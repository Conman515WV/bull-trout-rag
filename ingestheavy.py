"""
Yakima Fisheries RAG — Upgraded Ingestion Pipeline
====================================================
Parser:    Docling (PDF, DOCX, DOC) + openpyxl/xlrd (XLSX, XLS)
Chunking:  MarkdownHeaderTextSplitter → SemanticChunker
Embedding: BAAI/bge-large-en-v1.5
Summaries: Claude Haiku (optional, set SUMMARIZE = True)
Metadata:  Zotero SQLite lookup for publication-quality citations
Index:     ChromaDB (vectors) + BM25 (keyword)

Supported file types:
  .pdf, .pdf'  — Docling with OCR + table detection
  .docx, .doc  — Docling native
  .xlsx, .xls  — openpyxl / xlrd (sheet-by-sheet text extraction)
  .crdownload  — skipped (incomplete download)

Install dependencies before running:
    py -m pip install docling langchain-text-splitters langchain-experimental
    py -m pip install chromadb sentence-transformers anthropic rank-bm25
    py -m pip install openpyxl xlrd psutil

Runtime: ~6-10 hours for 400+ files on CPU
         Set SUMMARIZE = False to skip Haiku summaries and run ~2x faster
"""

import os
import re
import json
import pickle
import time
import sqlite3
import psutil
from pathlib import Path

# ── Low CPU priority ─────────────────────────────────────────────────────────
try:
    p = psutil.Process(os.getpid())
    p.nice(psutil.BELOW_NORMAL_PRIORITY_CLASS)
    print("Running at low CPU priority — your computer should stay responsive")
except Exception:
    pass

# ── Progress log ─────────────────────────────────────────────────────────────
LOG_FILE = "./ingest_progress.log"
with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write(f"Ingestion started: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

def log(msg):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(msg + "\n")

# ── Imports ───────────────────────────────────────────────────────────────────
from docling.document_converter import DocumentConverter
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import PdfFormatOption

from langchain_text_splitters import MarkdownHeaderTextSplitter
from langchain_experimental.text_splitter import SemanticChunker
from langchain_community.embeddings import HuggingFaceEmbeddings

from sentence_transformers import SentenceTransformer
from rank_bm25 import BM25Okapi
import chromadb
from anthropic import Anthropic

# ── Configuration ─────────────────────────────────────────────────────────────
SOURCE_FOLDER = r"C:\Users\Connor\Desktop\YakimaReferences"
CHROMA_PATH   = "./chroma_db"
BM25_PATH     = "./bm25_index.pkl"
REGISTRY_PATH = "./processed_files.json"
ZOTERO_DB     = r"C:\Users\Connor\Zotero\zotero.sqlite"

SUPPORTED_EXTENSIONS = {".pdf", ".pdf'", ".docx", ".doc", ".xlsx", ".xls"}
SKIP_EXTENSIONS      = {".crdownload"}

# Set to False to skip Haiku summaries — faster but slightly lower retrieval quality
SUMMARIZE = False

PARENT_SIZE = 1000
CHILD_SIZE  = 300

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# ── Zotero metadata lookup ────────────────────────────────────────────────────

def build_zotero_index(db_path):
    """
    Query Zotero SQLite to build a filename → metadata lookup dict.
    Returns dict keyed by PDF filename (lowercase, no extension).
    Closes the connection immediately so Zotero can still run.
    """
    index = {}
    if not os.path.exists(db_path):
        log(f"Zotero DB not found at {db_path} — falling back to heuristic titles")
        return index

    try:
        # Use a copy of the DB to avoid locking issues if Zotero is open
        import shutil, tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".sqlite", delete=False)
        tmp.close()
        shutil.copy2(db_path, tmp.name)

        conn = sqlite3.connect(tmp.name)
        cursor = conn.cursor()

        # Get title, year, authors, journal for each item with an attachment
        cursor.execute("""
            SELECT
                ia.itemID,
                title_val.value        AS title,
                date_val.value         AS date,
                publisher_val.value    AS journal,
                attach_path.path       AS attach_path
            FROM itemAttachments ia
            JOIN items attach_item ON ia.itemID = attach_item.itemID
            LEFT JOIN itemData title_d  ON ia.parentItemID = title_d.itemID  AND title_d.fieldID  = 1
            LEFT JOIN itemDataValues title_val ON title_d.valueID = title_val.valueID
            LEFT JOIN itemData date_d   ON ia.parentItemID = date_d.itemID   AND date_d.fieldID   = 6
            LEFT JOIN itemDataValues date_val  ON date_d.valueID  = date_val.valueID
            LEFT JOIN itemData pub_d    ON ia.parentItemID = pub_d.itemID    AND pub_d.fieldID    IN (12, 43)
            LEFT JOIN itemDataValues publisher_val ON pub_d.valueID = publisher_val.valueID
            LEFT JOIN (
                SELECT itemID, path FROM itemAttachments
            ) attach_path ON ia.itemID = attach_path.itemID
            WHERE ia.contentType = 'application/pdf'
               OR attach_item.itemTypeID = 3
        """)

        rows = cursor.fetchall()

        # Get authors separately
        cursor.execute("""
            SELECT
                ic.itemID,
                GROUP_CONCAT(
                    COALESCE(c.lastName, '') ||
                    CASE WHEN c.firstName != '' THEN ', ' || SUBSTR(c.firstName,1,1) || '.' ELSE '' END,
                    '; '
                ) AS authors
            FROM itemCreators ic
            JOIN creators c ON ic.creatorID = c.creatorID
            GROUP BY ic.itemID
        """)
        author_map = {row[0]: row[1] for row in cursor.fetchall()}

        conn.close()
        os.unlink(tmp.name)

        for item_id, title, date, journal, attach_path in rows:
            if not title:
                continue

            # Extract year from date string
            year = "unknown"
            if date:
                m = re.search(r'(19|20)\d{2}', str(date))
                if m:
                    year = m.group()

            # Get authors
            authors = author_map.get(item_id, "")

            # Build citation string
            citation = title
            if authors:
                # Get last name of first author
                first_author = authors.split(";")[0].split(",")[0].strip()
                et_al = " et al." if ";" in authors else ""
                citation = f"{first_author}{et_al} ({year}). {title}."
                if journal:
                    citation += f" {journal}."

            # Index by filename stem (lowercase) from attach path
            if attach_path:
                # Zotero stores paths like "storage:filename.pdf" or just filename
                fname = Path(attach_path.replace("storage:", "")).stem.lower()
                index[fname] = {
                    "title":    title,
                    "year":     year,
                    "authors":  authors,
                    "journal":  journal or "",
                    "citation": citation,
                }

        log(f"Zotero index built: {len(index)} entries")

    except Exception as e:
        log(f"Zotero lookup failed: {e} — falling back to heuristic titles")

    return index


def get_zotero_meta(filename, zotero_index):
    """
    Look up a file in the Zotero index by filename stem.
    Returns (title, year, authors, citation) tuple.
    Falls back to empty strings if not found.
    """
    stem = Path(filename).stem.lower()

    # Direct match
    if stem in zotero_index:
        m = zotero_index[stem]
        return m["title"], m["year"], m["authors"], m["citation"]

    # Fuzzy match — try stripping common suffixes like " (1)", " (2)"
    stem_clean = re.sub(r'\s*\(\d+\)\s*$', '', stem).strip()
    if stem_clean in zotero_index:
        m = zotero_index[stem_clean]
        return m["title"], m["year"], m["authors"], m["citation"]

    return "", "", "", ""

# ── Load models ───────────────────────────────────────────────────────────────
log("Loading embedding model (BAAI/bge-large-en-v1.5)...")
embed_model   = SentenceTransformer("BAAI/bge-large-en-v1.5")
lc_embeddings = HuggingFaceEmbeddings(model_name="BAAI/bge-large-en-v1.5")

anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY) if SUMMARIZE and ANTHROPIC_API_KEY else None

# ── Build Zotero metadata index ───────────────────────────────────────────────
log("Loading Zotero metadata...")
zotero_index = build_zotero_index(ZOTERO_DB)

# ── ChromaDB ──────────────────────────────────────────────────────────────────
chroma = chromadb.PersistentClient(path=CHROMA_PATH)
# Incremental ingestion: do NOT delete the collection. Get or create,
# and skip files already listed in processed_files.json.
collection = chroma.get_or_create_collection(name="yakima")
try:
    _existing_count = collection.count()
    log(f"Chroma collection opened: {_existing_count} existing chunks")
except Exception:
    pass

# ── Load processed-files registry (incremental ingestion) ────────────────
if os.path.exists(REGISTRY_PATH):
    with open(REGISTRY_PATH, "r", encoding="utf-8") as _f:
        processed_files = set(json.load(_f))
    log(f"Loaded registry: {len(processed_files)} already-indexed files will be skipped")
else:
    processed_files = set()
    log("No processed_files.json found — every file in SOURCE_FOLDER will be indexed.")
    log("Run bootstrap_registry.py first if your chroma_db already contains ingested files.")

# ── Docling converter ─────────────────────────────────────────────────────────
log("Setting up Docling converter...")
pipeline_options = PdfPipelineOptions()
pipeline_options.do_ocr             = True
pipeline_options.do_table_structure = True

converter = DocumentConverter(
    format_options={
        InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
    }
)

# ── Splitters ─────────────────────────────────────────────────────────────────
header_splitter = MarkdownHeaderTextSplitter(
    headers_to_split_on=[
        ("#",   "h1"),
        ("##",  "h2"),
        ("###", "h3"),
    ],
    strip_headers=False
)

semantic_splitter = SemanticChunker(
    lc_embeddings,
    breakpoint_threshold_type="percentile",
    breakpoint_threshold_amount=85
)

# ── Helper functions ──────────────────────────────────────────────────────────

def extract_year(filename, text):
    match = re.search(r'(19|20)\d{2}', filename)
    if match:
        return match.group()
    match = re.search(r'(19|20)\d{2}', text[:500])
    if match:
        return match.group()
    return "unknown"

def extract_title(text):
    lines = [l.strip() for l in text[:1000].split('\n') if len(l.strip()) > 20]
    return lines[0][:120] if lines else ""

def parse_excel(filepath):
    """Extract text from Excel files sheet by sheet."""
    ext = Path(filepath).suffix.lower()
    text_parts = []
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    text_parts.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))
            wb.close()
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for sheet in wb.sheets():
                rows = []
                for r in range(sheet.nrows):
                    row_text = "\t".join(str(sheet.cell_value(r, c)) for c in range(sheet.ncols))
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    text_parts.append(f"## Sheet: {sheet.name}\n" + "\n".join(rows))
    except Exception as e:
        log(f"    Excel parse error: {e}")
    return "\n\n".join(text_parts)

def parse_file(filepath):
    """Parse any supported file type to markdown text."""
    ext = Path(filepath).suffix.lower().rstrip("'")

    if ext in {".pdf", ".docx", ".doc"}:
        try:
            result = converter.convert(filepath)
            return result.document.export_to_markdown()
        except Exception as e:
            log(f"    Docling error: {e}")
            return ""

    elif ext in {".xlsx", ".xls"}:
        return parse_excel(filepath)

    return ""

def summarize_chunk(text):
    """Generate a concise summary using Claude Haiku."""
    if not anthropic_client:
        return ""
    try:
        response = anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=150,
            messages=[{"role": "user", "content": (
                f"Summarize this fisheries literature excerpt in 2-3 sentences. "
                f"Focus on key findings, species, location, and quantitative data. "
                f"Be specific and technical.\n\nExcerpt:\n{text[:1500]}\n\nSummary:"
            )}]
        )
        return response.content[0].text.strip()
    except Exception as e:
        log(f"    Summary failed: {e}")
        return ""

def chunk_text(markdown_text):
    """
    Two-stage chunking:
    1. Split by markdown headers
    2. Semantic chunking within each section
    Returns list of (parent_text, section_header) tuples
    """
    parents = []

    try:
        header_chunks = header_splitter.split_text(markdown_text)
    except Exception:
        header_chunks = [type('obj', (object,), {
            'page_content': markdown_text, 'metadata': {}
        })()]

    for hchunk in header_chunks:
        section_text = hchunk.page_content if hasattr(hchunk, 'page_content') else str(hchunk)
        header = ""
        if hasattr(hchunk, 'metadata') and hchunk.metadata:
            header = " > ".join(hchunk.metadata.values())

        if len(section_text.split()) < 50:
            if section_text.strip():
                parents.append((section_text.strip(), header))
            continue

        try:
            semantic_chunks = semantic_splitter.split_text(section_text)
            for chunk in semantic_chunks:
                if chunk.strip():
                    parents.append((chunk.strip(), header))
        except Exception:
            parents.append((section_text.strip(), header))

    return parents

# ── Collect files ─────────────────────────────────────────────────────────────
all_files     = []
skipped_files = []

for f in sorted(os.listdir(SOURCE_FOLDER)):
    ext = Path(f).suffix.lower()
    if ext in SUPPORTED_EXTENSIONS:
        all_files.append(f)
    else:
        skipped_files.append(f)

log(f"\nFound {len(all_files)} supported files")
log(f"Skipping {len(skipped_files)} unsupported/incomplete files")
log(f"Summarization: {'ON (Claude Haiku)' if SUMMARIZE else 'OFF'}\n")

ext_counts = {}
for f in all_files:
    ext = Path(f).suffix.lower()
    ext_counts[ext] = ext_counts.get(ext, 0) + 1
for ext, count in sorted(ext_counts.items()):
    log(f"  {count:4d}  {ext}")
log("")

# ── Ingestion loop ────────────────────────────────────────────────────────────
all_child_texts = []
all_child_ids   = []
all_child_metas = []
bm25_corpus     = []
bm25_texts      = []
bm25_metadata   = []

child_id  = 0
parent_id = 0
skipped   = 0
errors    = 0

start_time = time.time()

for i, filename in enumerate(all_files):
    # Skip files that were already indexed in a previous run.
    if filename in processed_files:
        log(f"[{i+1}/{len(all_files)}] {filename}  — already indexed, skipping")
        continue

    filepath = os.path.join(SOURCE_FOLDER, filename)
    elapsed  = time.time() - start_time
    eta      = (elapsed / max(i, 1)) * (len(all_files) - i)
    ext      = Path(filename).suffix.lower()
    log(f"[{i+1}/{len(all_files)}] {filename}  (ETA: {eta/60:.0f} min)")

    try:
        markdown_text = parse_file(filepath)

        if not markdown_text or len(markdown_text.strip()) < 100:
            log(f"  Skipping — no extractable text")
            skipped += 1
            continue

        year  = extract_year(filename, markdown_text)
        title = extract_title(markdown_text)
        ftype = ext.lstrip(".")

        # ── Zotero metadata lookup (overrides heuristics if found) ──────────
        z_title, z_year, z_authors, z_citation = get_zotero_meta(filename, zotero_index)
        if z_title:
            title  = z_title
            authors = z_authors
            citation = z_citation
            if z_year != "unknown":
                year = z_year
            log(f"  Zotero match: {citation[:80]}")
        else:
            authors  = ""
            citation = f"{title} ({year})" if title else filename

        parent_chunks = chunk_text(markdown_text)

        if not parent_chunks:
            log(f"  Skipping — no chunks produced")
            skipped += 1
            continue

        log(f"  {len(parent_chunks)} parent chunks  [{ftype}]")

        for parent_text, section_header in parent_chunks:
            pid = f"parent_{parent_id}"
            parent_id += 1

            summary    = summarize_chunk(parent_text) if SUMMARIZE else ""
            embed_text = summary if summary else parent_text

            tokens = parent_text.lower().split()
            bm25_corpus.append(tokens)
            bm25_texts.append(parent_text)
            bm25_metadata.append({
                "source":    filename,
                "year":      year,
                "title":     title,
                "authors":   authors,
                "citation":  citation,
                "section":   section_header,
                "summary":   summary,
                "filetype":  ftype,
                "parent_id": pid,
            })

            words = parent_text.split()
            step  = CHILD_SIZE - 30
            for start in range(0, len(words), step):
                chunk = " ".join(words[start:start + CHILD_SIZE])
                if not chunk.strip():
                    continue
                all_child_texts.append(embed_text)
                all_child_ids.append(str(child_id))
                all_child_metas.append({
                    "source":      filename,
                    "year":        year,
                    "title":       title,
                    "authors":     authors,
                    "citation":    citation,
                    "section":     section_header,
                    "summary":     summary,
                    "filetype":    ftype,
                    "parent_text": parent_text,
                    "parent_id":   pid,
                })
                child_id += 1

        # Mark file as successfully processed for incremental ingestion.
        processed_files.add(filename)
        with open(REGISTRY_PATH, "w", encoding="utf-8") as _rf:
            json.dump(sorted(processed_files), _rf, indent=2)

    except Exception as e:
        log(f"  ERROR: {e}")
        errors += 1
        continue

# ── Summary ───────────────────────────────────────────────────────────────────
log(f"\n{'='*60}")
log(f"Parsing complete:")
log(f"  Files processed: {len(all_files) - skipped - errors}")
log(f"  Skipped:         {skipped}")
log(f"  Errors:          {errors}")
log(f"  Parent chunks:   {parent_id}")
log(f"  Child chunks:    {child_id}")
log(f"{'='*60}\n")

# ── Embed ─────────────────────────────────────────────────────────────────────
log(f"Embedding {len(all_child_texts)} chunks with BAAI/bge-large-en-v1.5...")

embeddings = embed_model.encode(
    all_child_texts,
    show_progress_bar=True,
    batch_size=32,
    normalize_embeddings=True
)

# ── Store in ChromaDB ─────────────────────────────────────────────────────────
log("\nStoring in ChromaDB...")
batch_size = 200

for i in range(0, len(all_child_texts), batch_size):
    collection.add(
        ids=all_child_ids[i:i+batch_size],
        embeddings=embeddings[i:i+batch_size].tolist(),
        documents=all_child_texts[i:i+batch_size],
        metadatas=all_child_metas[i:i+batch_size]
    )
    log(f"  Stored {min(i+batch_size, len(all_child_texts))}/{len(all_child_texts)}")

# ── BM25 (incremental: load existing, append new, rebuild, save) ─────────────
log("\nBuilding BM25 keyword index (incremental merge with existing)...")

existing_corpus   = []
existing_texts    = []
existing_metadata = []
if os.path.exists(BM25_PATH):
    try:
        with open(BM25_PATH, "rb") as _bf:
            _old = pickle.load(_bf)
        existing_texts    = _old.get("texts", []) or []
        existing_metadata = _old.get("metadata", []) or []
        # Rebuild the tokenized corpus from the existing texts so the merged
        # BM25 scores stay consistent with the new material.
        existing_corpus = [t.lower().split() for t in existing_texts]
        log(f"Loaded existing BM25 index: {len(existing_texts)} parent chunks")
    except Exception as _e:
        log(f"Existing BM25 index unreadable ({_e}) — rebuilding from scratch.")
        existing_corpus, existing_texts, existing_metadata = [], [], []
else:
    log("No existing BM25 index — creating fresh one.")

merged_corpus   = existing_corpus   + bm25_corpus
merged_texts    = existing_texts    + bm25_texts
merged_metadata = existing_metadata + bm25_metadata

bm25 = BM25Okapi(merged_corpus) if merged_corpus else None

log(f"Saving BM25 index ({len(merged_texts)} parent chunks total)...")
with open(BM25_PATH, "wb") as f:
    pickle.dump({"bm25": bm25, "texts": merged_texts, "metadata": merged_metadata}, f)

# ── Done ──────────────────────────────────────────────────────────────────────
total_time = (time.time() - start_time) / 60
log(f"\n{'='*60}")
log(f"INGESTION COMPLETE in {total_time:.0f} minutes")
log(f"  ChromaDB:   {len(all_child_texts)} child chunks → {CHROMA_PATH}")
log(f"  BM25 index: {parent_id} parent chunks → {BM25_PATH}")
log(f"  Embed model: BAAI/bge-large-en-v1.5 (1024 dims)")
log(f"  Summarize:  {'ON' if SUMMARIZE else 'OFF'}")
log(f"{'='*60}")
log(f"\nNext steps:")
log(f"  1. Upload chroma_db/ and bm25_index.pkl to Hugging Face")
log(f"  2. Confirm app.py uses BAAI/bge-large-en-v1.5")
log(f"  3. Push app.py to GitHub and reboot Streamlit Cloud")